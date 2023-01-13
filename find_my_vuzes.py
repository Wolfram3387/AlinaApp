from openpyxl import load_workbook


def find_by_profession(profession):
    wb = load_workbook('VUZes.xlsx')
    sheet = wb['VUZes']
    result = {}
    for i in range(1, sheet.max_row + 1):
        if sheet[f'A{i}'].value == profession:
            result[i] = {}
            result[i]['Профессия'] = sheet[f'A{i}'].value
            result[i]['Университет'] = sheet[f'B{i}'].value
            result[i]['Факультет'] = sheet[f'C{i}'].value
            result[i]['Русский язык'] = sheet[f'D{i}'].value
            result[i]['Профильная математика'] = sheet[f'E{i}'].value
            result[i]['Физика'] = sheet[f'F{i}'].value
            result[i]['Информатика'] = sheet[f'G{i}'].value
            result[i]['Баллы на Бюджет'] = sheet[f'H{i}'].value
            result[i]['Сайт'] = sheet[f'I{i}'].value
    return result


def weed_out_disciplines(passed_disciplines, result):
    to_del = []
    for r in result:
        passed = [
            result[r]['Русский язык'],
            result[r]['Профильная математика'],
            result[r]['Физика'],
            result[r]['Информатика']
        ]
        if passed != passed_disciplines:
            to_del.append(r)
    for r_del in to_del:
        del result[r_del]
    return result


def weed_out_by_points(summa, result):
    to_del = []
    for r in result:
        min_points = result[r]['Баллы на Бюджет']
        if summa < min_points:
            to_del.append(r)
    for r_del in to_del:
        del result[r_del]
    return result


def find_my_VUZes(profession, disciplines, points):
    a1 = find_by_profession(profession)
    a2 = weed_out_disciplines(disciplines, a1)
    return weed_out_by_points(points, a2)
